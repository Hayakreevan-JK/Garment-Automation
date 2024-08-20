import time
import serial
import serial.tools.list_ports
import barcode as bc

from openpyxl import Workbook,load_workbook
arduinodata = serial.Serial('COM7',9600)
time.sleep(1)
wb = Workbook()
ws = wb.active
ws.title = 'Garments'
ws.append(['Barcode Scanner','Emp id','Op','Data Time'])
wb1 = load_workbook('pro2.xlsx')
ws1 = wb1.active
ws1.title = 'Barcode'
var = bc.x

while True:
    while(arduinodata.in_waiting == 0):
        pass
    datapacket = arduinodata.readline()
    datapacket = str(datapacket,'utf-8')
    datapacket = datapacket.strip('\r\n')
    print(datapacket)
    if(datapacket == 'E'):
        ws.append(x)
        x = []
        continue
    if(datapacket == '0'):
        break
    x.append(datapacket)

wb.save('pro1.xlsx')