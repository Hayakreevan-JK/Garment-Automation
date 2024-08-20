from openpyxl import Workbook,load_workbook
import serial.tools.list_ports
import time
import pandas as pd
import numpy as np
import serial
# import barcode as bc

# var = bc.x
# wb = load_workbook('pro1.xlsx')
# ws = wb.active
# excelfile = 'pro1.xlsx'
# df = pd.read_excel(excelfile)
# emp = int(input(""))
# op = int(input(""))
# fil1 = df["Emp id"] == emp
# fil2 = df["Op"] == op
# df2 = df[fil1 & fil2]
# res = df2.index[0]
# wb = Workbook()
# ws = wb.active
# ws.title = 'Garments'
# ws.append(['Index','Barcode'])
wb1 = load_workbook('pro2.xlsx')
ws1 = wb1.active
i = ws1.max_row
print(i)
ws1.append([123])

# while True:
#     i = int(input(""))
#     if(i == 0):
#         break
#     ws.append([res,i])
#     count += 1
# ws['C'+str(count)].value = count - 1

wb1.save('pro2.xlsx')
    



# wb1 = load_workbook('pro2.xlsx')
# ws1 = wb1.active
# ws1.title = 'Barcode'
# i = 2
# s = 'A'+str(i)
# while(ws1[s].value is None):
#     continue
# print("Hi")
